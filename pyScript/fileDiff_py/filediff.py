import openpyxl, os, string
from openpyxl.utils import get_column_letter, column_index_from_string

def getlen(sheet, col_str):
    len = 0
    begin = 4
    col = column_index_from_string(col_str)
    while True:
        cell = sheet.cell(begin + len, col)
        if not cell.value:
            break

        len += 1

    return len

def getCompNameList(sheet, col_str, len):
    list = []
    col = column_index_from_string(col_str)
    for i in range(len):
        cell = sheet.cell(i + 4, col)
        list.append(cell.value.strip())

    return list

def getnumbystr(cell):
    if not cell or cell.value == None:
        return 0

    return cell.value

def getfile(ans, filename):
    wb = openpyxl.load_workbook("ans.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    hight_row = sheet.max_row

    cell_title = sheet.cell(hight_row + 2, 1)
    cell_title.value = filename

    cnt = 0
    for k, v in ans.items():
        cell_name = sheet.cell(hight_row + 2 + cnt + 1, 1)
        cell_name.value = k

        cell_value = sheet.cell(hight_row + 2 + cnt + 1, 2)
        cell_value.value = v

        cnt += 1

    wb.save("ans.xlsx")


def removeSame(list1, list2):
    remove_list1 = []
    remove_list2 = []

    for v in list1:
        if not v in list2:
            remove_list1.append(v)

    for v in list2:
        if not v in list1:
            remove_list2.append(v)

    return remove_list1, remove_list2

def fileDiff(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name("应收账款")

    len1 = getlen(sheet, "B")
    len2 = getlen(sheet, "G")

    print(f"len1 = {len1}\tlen2 = {len2}")

    list1 = getCompNameList(sheet, "B", len1)
    list2 = getCompNameList(sheet, "G", len2)

    remove_list1, remove_list2 = removeSame(list1, list2)

    ans = {}

    col1 = column_index_from_string("B")
    col2 = column_index_from_string("G")
    for i in range(len1):

        isFind = False

        cell1 = sheet.cell(i + 4, col1)
        cell1_v1 = sheet.cell(i + 4, col1 + 1)
        cell1_v2 = sheet.cell(i + 4, col1 + 2)

        cell1_v1_dig = getnumbystr(cell1_v1)
        cell1_v2_dig = getnumbystr(cell1_v2)

        name1 = cell1.value.strip()
        for j in range(len2):
            cell2 = sheet.cell(j + 4, col2)
            cell2_v1 = sheet.cell(j + 4, col2 + 1)
            cell2_v2 = sheet.cell(j + 4, col2 + 2)

            cell2_v1_dig = getnumbystr(cell2_v1)
            cell2_v2_dig = getnumbystr(cell2_v2)
            name2 = cell2.value.strip()
            if(name1 == name2):
                if cell1_v1_dig != cell2_v1_dig or cell1_v2_dig != cell2_v2_dig:
                    diff = 0
                    if cell1_v1_dig != cell2_v1_dig:
                        diff = abs(cell1_v1_dig - cell2_v1_dig)
                    else:
                        diff = abs(cell1_v2_dig - cell2_v2_dig)

                    ans[name1] = diff
                    break
    print(ans)
    getfile(ans, filename)
