import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet import cell_range
import openpyxl.worksheet
import os

class excelManager:
    def read(self, file):
        if type(file) != str or file == "":
            print(f"type(file) = {type(file)}, file = {file}")
            raise AssertionError

        if not os.path.exists(file):
            print(f"not this file: {file}")
            raise AssertionError

        print(f"opening the file = {file}")

        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        if not sheet:
            print("not the active sheet!!!")
            return
        print(f"sheetName = {sheet.title}")
        print(f"max row = {sheet.max_row}, max col = {sheet.max_column}")
        print("输出每行数据")

        # print(sheet.merged_cells.ranges)
        merge_ranges = sheet.merged_cells.ranges
        mergeTuple = []
        for mergeRange in merge_ranges:
            for t in mergeRange.cells:
                mergeTuple.append(t)

        # for rows in sheet.rows:
        #     isvalue = False
        #     for cell in rows:
        #         if cell.value:
        #             isvalue = True
        #             print(f"{cell.coordinate}: {cell.value}", end= "\t")
        #
        #     if isvalue:
        #         print()
