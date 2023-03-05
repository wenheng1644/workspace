import openpyxl

import os

def excise1(filename, num):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("exicse1")

    for i in range(2, num + 2):
        rowCell = sheet.cell(1, i)
        rowCell.value = i-1

        colCell = sheet.cell(i, 1)
        colCell.value = i - 1


    for row in range(2, num + 2):
        for col in range(2, num + 2):
            cell = sheet.cell(row, col)
            preRowCell = sheet.cell(row,1)
            preColCell = sheet.cell(1, col)

            v1 = preRowCell.value
            v2 = preColCell.value

            cell.value = v1 * v2

            print(f"row = {row}, col = {col} = {v1*v2}")


    wb.save(filename)

def insertWithBack(filename, N, M = 1):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name("exicse1")
    sheet2 = wb.create_sheet(index = 0,title = sheet.title + "_copy")
    maxRow = sheet.max_row

    if N >= maxRow:
        print(f"don't need to do it, MaxRow = {maxRow}")
        return

    for row in range(1, maxRow+1):
        for col in range(1, sheet.max_column+1):
            cell = sheet.cell(row, col)

            transRow = row
            if row >= N:
                transRow += M

            if cell and cell.value:
                cell_cp = sheet2.cell(transRow, col)
                cell_cp.value = cell.value


    wb.save(filename)


def excise3(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name("Sheet_test")
    sheet2 = wb.create_sheet("convert_sheet")
    datas = []

    for rowCellObj in sheet.rows:
        rowData = []
        for cell in rowCellObj:
            rowData.append(cell.value)

        datas.append(rowData)

    print(datas)
    for x in range(len(datas)):
        for y in range(len(datas[x])):
            print(type(datas[x][y]), datas[x][y])
            cell = sheet2.cell(y+1, x +1)
            cell.value = datas[x][y]

    wb.save(filename)

