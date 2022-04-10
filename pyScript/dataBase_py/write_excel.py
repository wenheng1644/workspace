import os.path

import openpyxl,pprint
import src.products

from openpyxl.styles import Font,NamedStyle

def writeToExcel(filename):
    wb = openpyxl.load_workbook(filename)
    print(wb.get_sheet_names())
    sheet = wb.active
    print(f"old title: {sheet.title}")
    sheet.title = "test"
    print(wb.get_sheet_names())
    # wb.save("./src/write1.xlsx")
    mysheet = wb.create_sheet("first_sheet",0)
    print(wb.get_sheet_names())
    mysheet['A1'].value = "hello mygirl"
    mysheet['C3'].value = "刘烨仪"
    wb.save(filename)

def project_write(filename = None):
    proConf = src.products.products
    print(pprint.pformat(proConf))


    wb = openpyxl.load_workbook("./src/produceSales.xlsx")
    sheet = wb.get_sheet_by_name("Sheet")

    font = Font(size=24, italic=True)
    style_ = NamedStyle(font=font)
    for rowNum in range(2,sheet.max_row + 1):
        productName = sheet.cell(rowNum,1).value
        if productName in proConf:
            sheet.cell(rowNum,2).value = proConf[productName]
            sheet.cell(rowNum, 2).font = style_

    wb.save("./src/copy_product.xlsx")


def excise_excel(filename):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.title = "mylove"

    past = [["徐紫泳",22], ["刘烨仪", 23]]

    for row in range(1,len(past) + 1):
        for col in range(1,len(past[row-1]) + 1):
            sheet.cell(row,col).value = past[row-1][col-1]

            if past[row-1][col-1] == "徐紫泳":
                sheet.cell(row,col).font = Font(italic=True,name="Times New Roman",bold=True)


    # sheet.cell(3,2).value = "=SUM(B1:B2)"
    thecell = sheet.cell(1, 1).value
    if isinstance(thecell,str):
        len_text = len(thecell.encode("gbk"))
        print(f"中文长度： {len_text}")
    sheet.row_dimensions[1].height = 70
    sheet.column_dimensions['B'].width = 20


    wb.save(filename)
