import openpyxl

def excel_book(filename):
    wb = openpyxl.load_workbook(filename)
    sheets = wb.get_sheet_names()

    mysheet = wb.active
    print(mysheet['B1'].coordinate)
    print(mysheet.cell(2,2).value)

    rowCount = mysheet.max_row
    colCount = mysheet.max_column

    print(f"row: {rowCount}\tcol: {colCount}")
    # print(f"current activit sheet: {}")

    print(f"max letter: {openpyxl.utils.get_column_letter(colCount)}")
    print(f" to string : {openpyxl.utils.column_index_from_string('A')}")

def read_excelTable(filename, row, col, isAll = False):
    wb = openpyxl.load_workbook(filename)

    actSheet = wb.active

    print(f"the activeSheet: {actSheet.title}")

    if isAll:
        for rowInfo in actSheet.rows:
            for data in rowInfo:
                print(f"{data.coordinate}: {data.value}",end='\t')
            print()
    else:
        theCell = actSheet.cell(row,col)
        print(f"{theCell.coordinate}: {theCell.value}")


